import os
import pandas as pd
from openpyxl import load_workbook
import json

# Assign spreadsheet filename to `file`


def exportExcelTableAsJsonArray(fileName,importType,exportType):

    os.getcwd();

    file = str(fileName)+str(importType)

    wb = load_workbook("./files/excel/"+file)
    # Load spreadsheet
    sheet=wb.active

    max_row=sheet.max_row

    # get max column count
    max_column=sheet.max_column

    dataJson = []

    for i in range(1,max_row+1):
         jsonObject = {}
         # iterate over all columns
         for j in range(1,max_column+1):
              # get particular cell value
              cell_obj=sheet.cell(row=i,column=j)
              # print cell value
              jsonObject[sheet.cell(row=1,column=j).value] = cell_obj.value;

         # print new line
         #print(jsonObject,'\n')
         dataJson.append(jsonObject);

    with open('./files/json/'+fileName+exportType,'w') as outfile:
        json.dump(dataJson, outfile)


def exportExcelVertHorTableAsJsonArray(fileName,importType,exportType):

    os.getcwd();

    file = str(fileName)+str(importType)

    wb = load_workbook("./files/excel/"+file)
    # Load spreadsheet
    sheet=wb.active

    max_row=sheet.max_row

    # get max column count
    max_column=sheet.max_column

    dataJson = {}

    for i in range(2,max_row+1):
         jsonObject = {}

         # iterate over all columns
         for j in range(2,max_column+1):
              # get particular cell value
              cell_obj=sheet.cell(row=i,column=j)
              # print cell value
              jsonObject[sheet.cell(row=1,column=j).value] = cell_obj.value;

         # print new line
         print(jsonObject,'\n')
         dataJson[sheet.cell(row=i,column=1).value] = jsonObject

    with open('./files/json/'+fileName+exportType,'w') as outfile:
        json.dump(dataJson, outfile)

def readFilesInFolder(directory):
    jsonArray = []
    for root, dirs, files in os.walk(directory):
        jsonValue = []
        for filename in files:
            jsonValue = filename
            jsonArray.append(jsonValue)
    return jsonArray
